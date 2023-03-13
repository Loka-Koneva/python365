import openpyxl

from faker_data import gen_fake_data


def example():
    """
    Создание файла и запись в него
    """
    book = openpyxl.Workbook()
    book.remove(book.active)

    sheet_1 = book.create_sheet('Коллеги')
    sheet_2 = book.create_sheet('Клиенты')
    sheet_3 = book.create_sheet('Черный список', 0)

    for sheet in book.worksheets:
        for row in gen_fake_data():
            sheet.append(row)

    book.save('text.xlsx')

example()